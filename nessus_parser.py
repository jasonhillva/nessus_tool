#!/usr/bin/env python3
import os
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class NessusParser:
    def __init__(self, nessus_file=None):
        """Initialize with an optional .nessus file path"""
        self.nessus_file = nessus_file
        
    def parse(self):
        """
        Parse the .nessus file and return structured data
        
        Returns:
            dict: Dictionary containing scan info and vulnerability data
        """
        if not self.nessus_file:
            raise ValueError("No .nessus file specified")
            
        report_name, df = self.parse_nessus_file(self.nessus_file)
        
        # Extract scan information
        scan_info = {
            'Report Name': report_name,
            'File Path': self.nessus_file,
            'Date Processed': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Count vulnerabilities by severity
        if df is not None and not df.empty:
            severity_counts = df['Severity'].value_counts().to_dict()
            scan_info.update({
                'Critical': severity_counts.get('Critical', 0),
                'High': severity_counts.get('High', 0),
                'Medium': severity_counts.get('Medium', 0),
                'Low': severity_counts.get('Low', 0),
                'Info': severity_counts.get('Info', 0),
                'Total': len(df)
            })
        
        # Return structured data
        return {
            'scan_info': scan_info,
            'vulnerabilities': df.to_dict('records') if df is not None and not df.empty else []
        }

    @staticmethod
    def parse_nessus_file(nessus_file):
        """
        Parse a .nessus file and extract vulnerability information
        
        Args:
            nessus_file (str): Path to the .nessus file
            
        Returns:
            tuple: (scan_name, DataFrame with vulnerability data)
        """
        print(f"Parsing {nessus_file}...")
        
        try:
            tree = ET.parse(nessus_file)
            root = tree.getroot()
        except Exception as e:
            print(f"Error parsing {nessus_file}: {e}")
            return None, None
        
        # Extract scan metadata
        scan_metadata = {}
        
        # Get report name
        report_name = os.path.basename(nessus_file)
        report_elements = root.findall(".//Report")
        if report_elements and 'name' in report_elements[0].attrib:
            report_name = report_elements[0].attrib['name']
            scan_metadata['Report Name'] = report_name
        
        # Get policy name
        policy_elements = root.findall(".//Policy/policyName")
        if policy_elements and policy_elements[0].text:
            scan_metadata['Policy Name'] = policy_elements[0].text
        
        # Extract scan time and other preferences
        pref_elements = root.findall(".//preference")
        for pref in pref_elements:
            name_elem = pref.find("name")
            value_elem = pref.find("value")
            
            if name_elem is not None and value_elem is not None:
                name_text = name_elem.text
                value_text = value_elem.text
                
                if name_text == "report_host_details":
                    scan_metadata['Scan Time'] = value_text
                elif name_text == "report_paranoia":
                    scan_metadata['Report Paranoia'] = value_text
                elif name_text == "stop_scan_on_disconnect":
                    scan_metadata['Stop Scan on Disconnect'] = value_text
                elif name_text == "scan_start_time":
                    scan_metadata['Scan Start Time'] = value_text
                elif name_text == "scan_duration":
                    scan_metadata['Scan Duration'] = value_text
        
        # Extract target information from the policy (even if no hosts were found)
        target_list = []
        target_elements = root.findall(".//Preferences/ServerPreferences/preference")
        for pref in target_elements:
            name_elem = pref.find("name")
            if name_elem is not None and name_elem.text == "TARGET":
                value_elem = pref.find("value")
                if value_elem is not None:
                    targets = value_elem.text.split(",")
                    target_list.extend([t.strip() for t in targets])
                    scan_metadata['Targets'] = ", ".join([t.strip() for t in targets])
        
        # Extract host information and vulnerabilities
        vulnerabilities = []
        hosts_found = []
        
        # Find all ReportHost elements
        for host in root.findall(".//ReportHost"):
            host_name = host.attrib.get('name', 'Unknown')
            hosts_found.append(host_name)
            
            # Get IP, FQDN, and operating system if available
            ip_address = host_name
            fqdn = ""
            os_info = ""
            
            for tag in host.findall("./HostProperties/tag"):
                if tag.attrib.get('name') == "host-ip":
                    ip_address = tag.text
                elif tag.attrib.get('name') == "host-fqdn":
                    fqdn = tag.text
                elif tag.attrib.get('name') == "operating-system":
                    os_info = tag.text
            
            # Get all vulnerability items for this host
            for item in host.findall("./ReportItem"):
                plugin_id = item.attrib.get('pluginID', '')
                plugin_name = item.attrib.get('pluginName', '')
                port = item.attrib.get('port', '')
                protocol = item.attrib.get('protocol', '')
                severity = item.attrib.get('severity', '')
                
                # Convert severity number to text
                severity_text = {
                    '0': 'Info',
                    '1': 'Low',
                    '2': 'Medium',
                    '3': 'High',
                    '4': 'Critical'
                }.get(severity, severity)
                
                # Extract additional details
                description = ""
                solution = ""
                risk_factor = ""
                cvss_base_score = ""
                cvss3_base_score = ""
                cve = ""
                
                for child in item:
                    if child.tag == "description":
                        description = child.text or ""
                    elif child.tag == "solution":
                        solution = child.text or ""
                    elif child.tag == "risk_factor":
                        risk_factor = child.text or ""
                    elif child.tag == "cvss_base_score":
                        cvss_base_score = child.text or ""
                    elif child.tag == "cvss3_base_score":
                        cvss3_base_score = child.text or ""
                    elif child.tag == "cve":
                        cve = child.text or ""
                
                vulnerabilities.append({
                    'Host': host_name,
                    'IP Address': ip_address,
                    'FQDN': fqdn,
                    'Operating System': os_info,
                    'Port': port,
                    'Protocol': protocol,
                    'Plugin ID': plugin_id,
                    'Plugin Name': plugin_name,
                    'Severity': severity_text,
                    'Risk Factor': risk_factor,
                    'CVSS Base Score': cvss_base_score,
                    'CVSS3 Base Score': cvss3_base_score,
                    'CVE': cve,
                    'Description': description,
                    'Solution': solution
                })
        
        if hosts_found:
            scan_metadata['Hosts Found'] = ", ".join(hosts_found)
        
        # If no hosts were found but we have target information, create a placeholder entry
        if not vulnerabilities and target_list:
            print(f"No vulnerabilities found in {nessus_file}. Creating placeholder entry for targets: {target_list}")
            for target in target_list:
                entry = {
                    'Host': target,
                    'IP Address': target,
                    'FQDN': '',
                    'Operating System': '',
                    'Port': '',
                    'Protocol': '',
                    'Plugin ID': '',
                    'Plugin Name': 'No vulnerabilities found',
                    'Severity': 'Info',
                    'Risk Factor': '',
                    'CVSS Base Score': '',
                    'CVSS3 Base Score': '',
                    'CVE': '',
                    'Description': f'The scan did not find any vulnerabilities for target {target}.',
                    'Solution': ''
                }
                
                # Add scan metadata to entry
                for key, value in scan_metadata.items():
                    if key not in entry:  # Avoid overwriting existing keys
                        entry[key] = value
                
                vulnerabilities.append(entry)
        elif not vulnerabilities and hosts_found:
            print(f"Hosts found but no vulnerabilities in {nessus_file}. Creating placeholder entries for hosts: {hosts_found}")
            for host in hosts_found:
                entry = {
                    'Host': host,
                    'IP Address': host,
                    'FQDN': '',
                    'Operating System': '',
                    'Port': '',
                    'Protocol': '',
                    'Plugin ID': '',
                    'Plugin Name': 'No vulnerabilities found',
                    'Severity': 'Info',
                    'Risk Factor': '',
                    'CVSS Base Score': '',
                    'CVSS3 Base Score': '',
                    'CVE': '',
                    'Description': f'The scan did not find any vulnerabilities for host {host}.',
                    'Solution': ''
                }
                
                # Add scan metadata to entry
                for key, value in scan_metadata.items():
                    if key not in entry:  # Avoid overwriting existing keys
                        entry[key] = value
                
                vulnerabilities.append(entry)
        elif not vulnerabilities and not target_list and not hosts_found:
            print(f"No hosts or targets found in {nessus_file}. Creating a generic placeholder entry.")
            entry = {
                'Host': 'Unknown',
                'IP Address': 'Unknown',
                'FQDN': '',
                'Operating System': '',
                'Port': '',
                'Protocol': '',
                'Plugin ID': '',
                'Plugin Name': 'Empty scan results',
                'Severity': 'Info',
                'Risk Factor': '',
                'CVSS Base Score': '',
                'CVSS3 Base Score': '',
                'CVE': '',
                'Description': 'This scan did not contain any host or vulnerability data.',
                'Solution': ''
            }
            
            # Add scan metadata to entry
            for key, value in scan_metadata.items():
                if key not in entry:  # Avoid overwriting existing keys
                    entry[key] = value
            
            vulnerabilities.append(entry)
        
        # Convert to DataFrame
        if vulnerabilities:
            df = pd.DataFrame(vulnerabilities)
            
            # Sort by severity (Critical->High->Medium->Low->Info)
            severity_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3, 'Info': 4}
            df['SeverityOrder'] = df['Severity'].map(severity_order)
            df = df.sort_values('SeverityOrder').drop('SeverityOrder', axis=1)
            
            return report_name, df
        else:
            print(f"No vulnerability data found in {nessus_file}")
            return report_name, pd.DataFrame()

    @staticmethod
    def format_excel(writer, sheet_name):
        """
        Format the Excel worksheet for better readability
        
        Args:
            writer (pd.ExcelWriter): Excel writer object
            sheet_name (str): Name of the sheet to format
        """
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Define formats for different severity levels
        severity_colors = {
            'Critical': 'FF0000',  # Red
            'High': 'FFA500',      # Orange
            'Medium': 'FFFF00',    # Yellow
            'Low': 'ADFF2F',       # Green Yellow
            'Info': '00FFFF'       # Cyan
        }
        
        # Get the column index for the Severity column
        severity_col = None
        for i, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1)):
            if col[0].value == 'Severity':
                severity_col = i + 1
                break
        
        if severity_col:
            # Apply conditional formatting based on severity
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=severity_col)
                if cell.value in severity_colors:
                    color = severity_colors[cell.value]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Format header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Adjust column widths
        for i, column in enumerate(worksheet.columns, 1):
            column_width = 15  # Default width
            
            # Make some columns wider
            if column[0].value in ['Plugin Name', 'Description', 'Solution']:
                column_width = 60
            elif column[0].value in ['Host', 'IP Address', 'FQDN', 'Operating System', 'CVE']:
                column_width = 25
                
            worksheet.column_dimensions[get_column_letter(i)].width = column_width
        
        # Wrap text in description and solution columns
        wrap_columns = ['Plugin Name', 'Description', 'Solution']
        for col_name in wrap_columns:
            col_idx = None
            for i, col in enumerate(worksheet.iter_cols(min_row=1, max_row=1)):
                if col[0].value == col_name:
                    col_idx = i + 1
                    break
                    
            if col_idx:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.alignment = Alignment(wrapText=True, vertical='top')
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'

    @staticmethod
    def export_to_excel(nessus_files, output_file):
        """
        Parse multiple .nessus files and export the data to an Excel file with multiple tabs
        
        Args:
            nessus_files (list): List of paths to .nessus files
            output_file (str): Path to the output Excel file
        """
        if not nessus_files:
            print("No .nessus files specified!")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Process each nessus file
            summary_data = []
            has_valid_data = False
            
            for nessus_file in nessus_files:
                try:
                    report_name, df = NessusParser.parse_nessus_file(nessus_file)
                    
                    if df is not None and not df.empty:
                        has_valid_data = True
                        
                        # Create valid sheet name (Excel has a 31 character limit and doesn't allow certain characters)
                        sheet_name = report_name
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:28] + "..."
                        
                        # Replace invalid characters
                        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
                        for char in invalid_chars:
                            sheet_name = sheet_name.replace(char, '_')
                        
                        # Check for duplicate sheet names
                        sheet_count = 1
                        original_name = sheet_name
                        while sheet_name in writer.sheets:
                            suffix = f"_{sheet_count}"
                            sheet_name = original_name[:31-len(suffix)] + suffix
                            sheet_count += 1
                        
                        # Write data to Excel
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Format the sheet
                        NessusParser.format_excel(writer, sheet_name)
                        
                        # Collect summary information
                        severity_counts = df['Severity'].value_counts().to_dict()
                        
                        summary_data.append({
                            'Report Name': report_name,
                            'Critical': severity_counts.get('Critical', 0),
                            'High': severity_counts.get('High', 0),
                            'Medium': severity_counts.get('Medium', 0),
                            'Low': severity_counts.get('Low', 0),
                            'Info': severity_counts.get('Info', 0),
                            'Total': len(df)
                        })
                except Exception as e:
                    print(f"Error processing {nessus_file}: {e}")
            
            # Create summary sheet if we have data
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                summary_df = summary_df[['Report Name', 'Critical', 'High', 'Medium', 'Low', 'Info', 'Total']]
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                NessusParser.format_excel(writer, 'Summary')
                
                # Make the Summary sheet the active sheet when opening the file
                workbook = writer.book
                workbook.active = 0
            elif not has_valid_data:
                # If no data was processed successfully, create an empty sheet to prevent Excel error
                pd.DataFrame().to_excel(writer, sheet_name='No Data')
        
        print(f"Exported vulnerabilities to {output_file}")

    @staticmethod
    def export_to_csv(nessus_files, output_dir):
        """
        Parse multiple .nessus files and export the data to CSV files
        
        Args:
            nessus_files (list): List of paths to .nessus files
            output_dir (str): Directory to save CSV files
        """
        if not nessus_files:
            print("No .nessus files specified!")
            return
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Process each nessus file
        summary_data = []
        
        for nessus_file in nessus_files:
            try:
                report_name, df = NessusParser.parse_nessus_file(nessus_file)
                
                if df is not None and not df.empty:
                    # Create valid filename
                    safe_name = "".join(c if c.isalnum() or c in ['-', '_', '.'] else '_' for c in report_name)
                    csv_file = os.path.join(output_dir, f"{safe_name}.csv")
                    
                    # Write data to CSV
                    df.to_csv(csv_file, index=False)
                    print(f"Exported vulnerabilities to {csv_file}")
                    
                    # Collect summary information
                    severity_counts = df['Severity'].value_counts().to_dict()
                    
                    summary_data.append({
                        'Report Name': report_name,
                        'Critical': severity_counts.get('Critical', 0),
                        'High': severity_counts.get('High', 0),
                        'Medium': severity_counts.get('Medium', 0),
                        'Low': severity_counts.get('Low', 0),
                        'Info': severity_counts.get('Info', 0),
                        'Total': len(df)
                    })
            except Exception as e:
                print(f"Error processing {nessus_file}: {e}")
        
        # Create summary CSV if we have data
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df = summary_df[['Report Name', 'Critical', 'High', 'Medium', 'Low', 'Info', 'Total']]
            summary_csv = os.path.join(output_dir, "summary.csv")
            summary_df.to_csv(summary_csv, index=False)
            print(f"Exported summary to {summary_csv}")